from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from app.models import Course, Session_Year, CustomUser, Student, Staff, Subject, Staff_Notification, Staff_leave, Staff_Feedback, Student_Notification, Student_Feedback, Student_leave, Attendance, Attendance_Report, StudentResult, Note
from django.contrib import messages
import openpyxl

@login_required(login_url='/')
def HOME(request):
    staff = Staff.objects.get(admin=request.user)
    leave_count = Staff_leave.objects.filter(staff_id=staff, status=0).count()
    feedback_count = Staff_Feedback.objects.filter(staff_id=staff, feedback_reply="").count()
    subject_count = staff.subjects.count()
    notification_count = Staff_Notification.objects.filter(staff_id=staff, status=0).count()
    notes_count = Note.objects.filter(user=request.user).count()

    context = {
        'leave_count': leave_count,
        'feedback_count': feedback_count,
        'subject_count': subject_count,
        'notification_count': notification_count,
        'notes_count': notes_count,
    }
    return render(request, 'Staff/home.html', context)

@login_required(login_url='/')
def NOTIFICATIONS(request):
    staff = Staff.objects.filter(admin=request.user)
    for i in staff:
        staff_id = i.id
        notification = Staff_Notification.objects.filter(staff_id=staff_id)
        context = {
            'notification': notification,
        }
        return render(request, 'Staff/notification.html', context)

@login_required(login_url='/')
def STAFF_NOTIFICATION_MARK_AS_DONE(request, status):
    notification = Staff_Notification.objects.get(id=status)
    notification.status = 1
    notification.save()
    return redirect('notifications')

@login_required(login_url='/')
def STAFF_APPLY_LEAVE(request):
    staff = Staff.objects.filter(admin=request.user)
    for i in staff:
        staff_id = i.id
        staff_leave_history = Staff_leave.objects.filter(staff_id=staff_id)
        context = {
            'staff_leave_history': staff_leave_history,
        }
        return render(request, 'Staff/apply_leave.html', context)

@login_required(login_url='/')
def STAFF_APPLY_LEAVE_SAVE(request):
    if request.method == "POST":
        leave_date = request.POST.get('leave_date')
        leave_message = request.POST.get('message')

        if not leave_date:
            messages.error(request, 'Please provide a valid leave date.')
            return redirect('staff_apply_leave')
        if not leave_message or leave_message.strip() == '':
            messages.error(request, 'Please provide a reason for the leave.')
            return redirect('staff_apply_leave')

        staff = Staff.objects.get(admin=request.user)
        staff_leave = Staff_leave(
            staff_id=staff,
            date=leave_date,
            message=leave_message,
        )
        staff_leave.save()
        messages.success(request, 'Leave Application Successfully Sent!')
        return redirect('staff_apply_leave')
    return redirect('staff_apply_leave')

@login_required(login_url='/')
def STAFF_FEEDBACK(request):
    staff = Staff.objects.get(admin=request.user)
    feedback_history = Staff_Feedback.objects.filter(staff_id=staff).order_by('-id')
    context = {
        'feedback_history': feedback_history,
    }
    return render(request, 'Staff/feedback.html', context)

@login_required(login_url='/')
def STAFF_SEND_FEEDBACK(request):
    if request.method == "POST":
        feedback = request.POST.get('feedback')
        staff = Staff.objects.get(admin=request.user)
        feedbacks = Staff_Feedback(
            staff_id=staff,
            feedback=feedback,
            feedback_reply="",
        )
        feedbacks.save()
        messages.success(request, 'Feedback Successfully Sent!')
        return redirect('staff_feedback')
    return redirect('staff_feedback')

@login_required(login_url='/')
def STAFF_TAKE_ATTENDANCE(request):
    staff = Staff.objects.get(admin=request.user)
    subject = staff.subjects.all()
    session_year = Session_Year.objects.all()
    action = request.GET.get('action')
    get_subject = None
    get_session_year = None
    students = None

    if action == 'get_student':
        if request.method == "POST":
            subject_id = request.POST.get('subject_id')
            session_year_id = request.POST.get('session_year_id')
            try:
                get_subject = Subject.objects.get(id=subject_id)
                get_session_year = Session_Year.objects.get(id=session_year_id)
                if get_subject not in staff.subjects.all():
                    messages.error(request, "You are not authorized to take attendance for this subject.")
                    return redirect('staff_take_attendance')
                students = Student.objects.filter(course_id=get_subject.course, session_year_id=get_session_year)
            except Subject.DoesNotExist:
                messages.error(request, "Selected subject does not exist.")
                return redirect('staff_take_attendance')
            except Session_Year.DoesNotExist:
                messages.error(request, "Selected session year does not exist.")
                return redirect('staff_take_attendance')

    context = {
        'subject': subject,
        'session_year': session_year,
        'action': action,
        'get_subject': get_subject,
        'get_session_year': get_session_year,
        'students': students,
    }
    return render(request, 'Staff/take_attendance.html', context)

@login_required(login_url='/')
def STAFF_SAVE_ATTENDANCE(request):
    if request.method == "POST":
        subject_id = request.POST.get('subject_id')
        session_year_id = request.POST.get('session_year_id')
        attendance_date = request.POST.get('attendance_date')
        student_ids = request.POST.getlist('student_id')

        try:
            get_subject = Subject.objects.get(id=subject_id)
            get_session_year = Session_Year.objects.get(id=session_year_id)
        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")
            return redirect('staff_take_attendance')
        except Session_Year.DoesNotExist:
            messages.error(request, "Selected session year does not exist.")
            return redirect('staff_take_attendance')

        staff = Staff.objects.get(admin=request.user)
        if get_subject not in staff.subjects.all():
            messages.error(request, "You are not authorized to take attendance for this subject.")
            return redirect('staff_take_attendance')

        attendance, created = Attendance.objects.get_or_create(
            subject_id=get_subject,
            attendance_data=attendance_date,
            session_year_id=get_session_year
        )

        Attendance_Report.objects.filter(attendance_id=attendance).delete()

        students = Student.objects.filter(course_id=get_subject.course, session_year_id=get_session_year)
        present_student_ids = set(student_ids)

        for student in students:
            status = 1 if str(student.id) in present_student_ids else 0
            attendance_report = Attendance_Report(
                student_id=student,
                attendance_id=attendance,
                status=status
            )
            attendance_report.save()

        messages.success(request, "Attendance saved successfully.")
        redirect_url = reverse('staff_view_attendance') + f'?action=view_attendance&subject_id={subject_id}&session_year_id={session_year_id}&attendance_date={attendance_date}'
        return redirect(redirect_url)

    return redirect('staff_take_attendance')

@login_required(login_url='/')
def STAFF_VIEW_ATTENDANCE(request):
    staff = Staff.objects.get(admin=request.user)
    subjects = staff.subjects.all()
    session_years = Session_Year.objects.all()
    action = request.GET.get('action')
    get_subject = None
    get_session_year = None
    start_date = None
    end_date = None
    attendance_reports = None

    if action == 'view_attendance' or (request.method == "GET" and 'subject_id' in request.GET):
        if request.method == "POST":
            subject_id = request.POST.get('subject_id')
            session_year_id = request.POST.get('session_year_id')
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
        else:
            subject_id = request.GET.get('subject_id')
            session_year_id = request.GET.get('session_year_id')
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

        try:
            get_subject = Subject.objects.get(id=subject_id)
            get_session_year = Session_Year.objects.get(id=session_year_id)
            
            if get_subject not in staff.subjects.all():
                messages.error(request, "You are not authorized to view attendance for this subject.")
                return redirect('staff_view_attendance')
                
            attendance_query = Attendance.objects.filter(
                subject_id=get_subject,
                session_year_id=get_session_year
            )
            
            if start_date:
                attendance_query = attendance_query.filter(attendance_data__gte=start_date)
            if end_date:
                attendance_query = attendance_query.filter(attendance_data__lte=end_date)
                
            attendance_reports = Attendance_Report.objects.filter(
                attendance_id__in=attendance_query
            ).order_by('attendance_id__attendance_data', 'student_id__admin__first_name')

        except Subject.DoesNotExist:
            messages.error(request, "Selected subject does not exist.")
            return redirect('staff_view_attendance')
        except Session_Year.DoesNotExist:
            messages.error(request, "Selected session year does not exist.")
            return redirect('staff_view_attendance')

    if 'download' in request.GET and request.GET['download'] == 'excel':
        subject_id = request.GET.get('subject_id')
        session_year_id = request.GET.get('session_year_id')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        try:
            get_subject = Subject.objects.get(id=subject_id)
            get_session_year = Session_Year.objects.get(id=session_year_id)
            
            if get_subject not in staff.subjects.all():
                messages.error(request, "You are not authorized to download attendance for this subject.")
                return redirect('staff_view_attendance')
                
            attendance_query = Attendance.objects.filter(
                subject_id=get_subject,
                session_year_id=get_session_year
            )
            
            if start_date:
                attendance_query = attendance_query.filter(attendance_data__gte=start_date)
            if end_date:
                attendance_query = attendance_query.filter(attendance_data__lte=end_date)
                
            attendance_reports = Attendance_Report.objects.filter(
                attendance_id__in=attendance_query
            ).order_by('attendance_id__attendance_data', 'student_id__admin__first_name')

            if attendance_reports.exists():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Attendance"
                ws.append(["Student ID", "Student Name", "Subject", "Date", "Status"])
                for report in attendance_reports:
                    student = report.student_id
                    ws.append([
                        student.admin.id,
                        f"{student.admin.first_name} {student.admin.last_name}",
                        report.attendance_id.subject_id.name,
                        report.attendance_id.attendance_data.strftime('%Y-%m-%d'),
                        "Present" if report.status == 1 else "Absent"
                    ])
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename=attendance_{get_subject.name}_{get_session_year.session_start}_to_{get_session_year.session_end}.xlsx'
                wb.save(response)
                return response
            else:
                messages.error(request, "No attendance data available to export.")
                return redirect('staff_view_attendance')

        except (Subject.DoesNotExist, Session_Year.DoesNotExist):
            messages.error(request, "Invalid subject or session year.")
            return redirect('staff_view_attendance')

    context = {
        'subjects': subjects,
        'session_years': session_years,
        'action': action,
        'get_subject': get_subject,
        'get_session_year': get_session_year,
        'start_date': start_date,
        'end_date': end_date,
        'attendance_reports': attendance_reports,
    }
    return render(request, 'Staff/view_attendance.html', context)

@login_required(login_url='/')
def STAFF_ADD_RESULT(request):
    staff = Staff.objects.get(admin=request.user)
    subjects = staff.subjects.all()
    session_year = Session_Year.objects.all()
    action = request.GET.get('action')
    get_subject = None
    get_session = None
    students = None
    results = None
    success_message = None

    if 'download' in request.GET and request.GET['download'] == 'excel':
        subject_id = request.GET.get('subject_id')
        session_year_id = request.GET.get('session_year_id')

        try:
            get_subject = Subject.objects.get(id=subject_id)
            get_session = Session_Year.objects.get(id=session_year_id)
            if get_subject not in staff.subjects.all():
                messages.error(request, "You are not authorized to download results for this subject.")
                return redirect('staff_add_result')
            students = Student.objects.filter(course_id=get_subject.course, session_year_id=get_session)
            results = StudentResult.objects.filter(subject_id=get_subject, student_id__session_year_id=get_session)
        except (Subject.DoesNotExist, Session_Year.DoesNotExist, ValueError) as e:
            messages.error(request, "Invalid subject or session year selected.")
            return redirect('staff_add_result')

        if students.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(["Student Name", "Student ID", "Assignment", "Exam", "IA1", "IA2", "Attendance", "Mid Sem", "End Sem", "Total", "CGPA"])
            for student in students:
                result = results.filter(student_id=student).first()
                ws.append([
                    f"{student.admin.first_name} {student.admin.last_name}",
                    student.id,
                    result.assignment_mark if result and result.assignment_mark is not None else '-',
                    result.exam_mark if result and result.exam_mark is not None else '-',
                    result.ia1_mark if result and result.ia1_mark is not None else '-',
                    result.ia2_mark if result and result.ia2_mark is not None else '-',
                    result.attendance_mark if result and result.attendance_mark is not None else '-',
                    result.midsem_mark if result and result.midsem_mark is not None else '-',
                    result.end_sem_mark if result and result.end_sem_mark is not None else '-',
                    result.total_mark if result and result.total_mark is not None else '-',
                    result.calculate_cgpa() if result and result.calculate_cgpa() is not None else '-'
                ])
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=results_{get_subject.name}_{get_session.session_start}_to_{get_session.session_end}.xlsx'
            wb.save(response)
            return response
        else:
            messages.error(request, "No results data available to export.")
            return redirect('staff_add_result')

    if action == "get_student":
        if request.method == "POST":
            subject_id = request.POST.get('subject_id')
            session_year_id = request.POST.get('session_year_id')
            excel_file = request.FILES.get('excel_file')

            try:
                get_subject = Subject.objects.get(id=subject_id)
                get_session = Session_Year.objects.get(id=session_year_id)
                if get_subject not in staff.subjects.all():
                    messages.error(request, "You are not authorized to add results for this subject.")
                    return redirect('staff_add_result')
                students = Student.objects.filter(course_id=get_subject.course, session_year_id=get_session)
                results = StudentResult.objects.filter(subject_id=get_subject, student_id__session_year_id=get_session)
            except Subject.DoesNotExist:
                messages.error(request, "Selected subject does not exist.")
                return redirect('staff_add_result')
            except Session_Year.DoesNotExist:
                messages.error(request, "Selected session year does not exist.")
                return redirect('staff_add_result')

            # Handle Excel upload
            if excel_file:
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    header_map = {cell.lower().strip(): idx for idx, cell in enumerate(header_row) if cell}

                    required_columns = ['student id', 'assignment', 'exam', 'ia1', 'ia2', 'attendance', 'mid sem', 'end sem']
                    missing_columns = [col for col in required_columns if col not in header_map]

                    if missing_columns:
                        messages.error(request, f"Missing columns in Excel file: {', '.join(missing_columns)}")
                        return redirect('staff_add_result')

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        student_id = row[header_map['student id']]
                        assignment_mark = row[header_map['assignment']]
                        exam_mark = row[header_map['exam']]
                        ia1_mark = row[header_map['ia1']]
                        ia2_mark = row[header_map['ia2']]
                        attendance_mark = row[header_map['attendance']]
                        midsem_mark = row[header_map['mid sem']]
                        end_sem_mark = row[header_map['end sem']]

                        try:
                            student_id = int(student_id)
                            assignment_mark = int(assignment_mark) if assignment_mark else None
                            exam_mark = int(exam_mark) if exam_mark else None
                            ia1_mark = float(ia1_mark) if ia1_mark else None
                            ia2_mark = float(ia2_mark) if ia2_mark else None
                            attendance_mark = float(attendance_mark) if attendance_mark else None
                            midsem_mark = float(midsem_mark) if midsem_mark else None
                            end_sem_mark = float(end_sem_mark) if end_sem_mark else None
                        except (ValueError, TypeError):
                            messages.error(request, f"Invalid data in row for Student ID {student_id}")
                            continue

                        try:
                            student = Student.objects.get(id=student_id, course_id=get_subject.course, session_year_id=get_session)
                        except Student.DoesNotExist:
                            messages.error(request, f"Student with ID {student_id} not found for this subject and session year.")
                            continue

                        # Update or create the result, ensuring total_mark is recalculated
                        result, created = StudentResult.objects.get_or_create(
                            student_id=student,
                            subject_id=get_subject,
                            defaults={
                                'assignment_mark': assignment_mark,
                                'exam_mark': exam_mark,
                                'ia1_mark': ia1_mark,
                                'ia2_mark': ia2_mark,
                                'attendance_mark': attendance_mark,
                                'midsem_mark': midsem_mark,
                                'end_sem_mark': end_sem_mark,
                            }
                        )
                        if not created:
                            # Update existing result and trigger total_mark recalculation
                            result.assignment_mark = assignment_mark
                            result.exam_mark = exam_mark
                            result.ia1_mark = ia1_mark
                            result.ia2_mark = ia2_mark
                            result.attendance_mark = attendance_mark
                            result.midsem_mark = midsem_mark
                            result.end_sem_mark = end_sem_mark
                            result.save()  # Triggers calculate_total_marks via save method
                    success_message = "Results uploaded successfully."
                except Exception as e:
                    messages.error(request, f"Error processing file: {str(e)}")
                    return redirect('staff_add_result')

            # Handle manual input
            elif 'save_all' in request.POST:
                for student in students:
                    assignment_mark = request.POST.get(f'assignment_mark_{student.id}')
                    exam_mark = request.POST.get(f'exam_mark_{student.id}')
                    ia1_mark = request.POST.get(f'ia1_mark_{student.id}')
                    ia2_mark = request.POST.get(f'ia2_mark_{student.id}')
                    attendance_mark = request.POST.get(f'attendance_mark_{student.id}')
                    midsem_mark = request.POST.get(f'midsem_mark_{student.id}')
                    end_sem_mark = request.POST.get(f'end_sem_mark_{student.id}')

                    if not all([assignment_mark, exam_mark, ia1_mark, ia2_mark, attendance_mark, midsem_mark, end_sem_mark]):
                        messages.error(request, f"All fields must be filled for student {student.admin.first_name} {student.admin.last_name}.")
                        continue

                    try:
                        assignment_mark = int(assignment_mark)
                        exam_mark = int(exam_mark)
                        ia1_mark = float(ia1_mark)
                        ia2_mark = float(ia2_mark)
                        attendance_mark = float(attendance_mark)
                        midsem_mark = float(midsem_mark)
                        end_sem_mark = float(end_sem_mark)
                    except (ValueError, TypeError):
                        messages.error(request, f"Invalid data for student {student.admin.first_name} {student.admin.last_name}.")
                        continue

                    # Update or create the result, ensuring total_mark is recalculated
                    result, created = StudentResult.objects.get_or_create(
                        student_id=student,
                        subject_id=get_subject,
                        defaults={
                            'assignment_mark': assignment_mark,
                            'exam_mark': exam_mark,
                            'ia1_mark': ia1_mark,
                            'ia2_mark': ia2_mark,
                            'attendance_mark': attendance_mark,
                            'midsem_mark': midsem_mark,
                            'end_sem_mark': end_sem_mark,
                        }
                    )
                    if not created:
                        # Update existing result and trigger total_mark recalculation
                        result.assignment_mark = assignment_mark
                        result.exam_mark = exam_mark
                        result.ia1_mark = ia1_mark
                        result.ia2_mark = ia2_mark
                        result.attendance_mark = attendance_mark
                        result.midsem_mark = midsem_mark
                        result.end_sem_mark = end_sem_mark
                        result.save()  # Triggers calculate_total_marks via save method
                    if result.id:
                        success_message = "Results saved successfully."

    context = {
        'subjects': subjects,
        'session_year': session_year,
        'action': action,
        'get_subject': get_subject,
        'get_session': get_session,
        'students': students,
        'results': results,
        'success_message': success_message,
    }
    return render(request, 'Staff/add_result.html', context)

@login_required(login_url='/')
def STAFF_SAVE_RESULT(request):
    return redirect('staff_add_result')

@login_required(login_url='/')
def STAFF_VIEW_RESULT(request):
    staff = Staff.objects.get(admin=request.user)
    subjects = staff.subjects.all()
    session_year = Session_Year.objects.all()
    action = request.GET.get('action')
    get_subject = None
    get_session = None
    students = None
    results = None

    if action == "view_result":
        if request.method == "POST":
            subject_id = request.POST.get('subject_id')
            session_year_id = request.POST.get('session_year_id')
            try:
                get_subject = Subject.objects.get(id=subject_id)
                get_session = Session_Year.objects.get(id=session_year_id)
                if get_subject not in staff.subjects.all():
                    messages.error(request, "You are not authorized to view results for this subject.")
                    return redirect('staff_view_result')
                students = Student.objects.filter(course_id=get_subject.course, session_year_id=get_session)
                results = StudentResult.objects.filter(subject_id=get_subject, student_id__session_year_id=get_session)
            except Subject.DoesNotExist:
                messages.error(request, "Selected subject does not exist.")
                return redirect('staff_view_result')
            except Session_Year.DoesNotExist:
                messages.error(request, "Selected session year does not exist.")
                return redirect('staff_view_result')

    if 'download' in request.GET and request.GET['download'] == 'excel' and get_subject and get_session:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["Student Name", "Assignment", "Exam", "IA1", "IA2", "Attendance", "Mid Sem", "End Sem", "Total", "CGPA"])
        for student in students:
            result = results.filter(student_id=student).first()
            ws.append([
                f"{student.admin.first_name} {student.admin.last_name}",
                result.assignment_mark if result and result.assignment_mark is not None else '-',
                result.exam_mark if result and result.exam_mark is not None else '-',
                result.ia1_mark if result and result.ia1_mark is not None else '-',
                result.ia2_mark if result and result.ia2_mark is not None else '-',
                result.attendance_mark if result and result.attendance_mark is not None else '-',
                result.midsem_mark if result and result.midsem_mark is not None else '-',
                result.end_sem_mark if result and result.end_sem_mark is not None else '-',
                result.total_mark if result and result.total_mark is not None else '-',
                result.calculate_cgpa() if result and result.calculate_cgpa() is not None else '-'
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=results_{get_subject.name}_{get_session.session_start}_to_{get_session.session_end}.xlsx'
        wb.save(response)
        return response

    context = {
        'subjects': subjects,
        'session_year': session_year,
        'action': action,
        'get_subject': get_subject,
        'get_session': get_session,
        'students': students,
        'results': results,
    }
    return render(request, 'Staff/view_result.html', context)

@login_required(login_url='/')
def STAFF_DOWNLOAD_RESULT_TEMPLATE(request):
    subject_id = request.GET.get('subject_id')
    session_year_id = request.GET.get('session_year_id')

    if not all([subject_id, session_year_id]):
        messages.error(request, "Please select subject and session year.")
        return redirect('staff_add_result')

    try:
        get_subject = Subject.objects.get(id=subject_id)
        get_session = Session_Year.objects.get(id=session_year_id)
        staff = Staff.objects.get(admin=request.user)
        if get_subject not in staff.subjects.all():
            messages.error(request, "You are not authorized to download template for this subject.")
            return redirect('staff_add_result')
        students = Student.objects.filter(course_id=get_subject.course, session_year_id=get_session)
    except (Subject.DoesNotExist, Session_Year.DoesNotExist):
        messages.error(request, "Invalid subject or session year selected.")
        return redirect('staff_add_result')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results Template"
    ws.append(["Student ID", "Assignment", "Exam", "IA1", "IA2", "Attendance", "Mid Sem", "End Sem"])
    for student in students:
        ws.append([student.id, "", "", "", "", "", "", ""])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=result_template_{get_subject.name}_{get_session.session_start}_to_{get_session.session_end}.xlsx'
    wb.save(response)
    return response

# New views for staff notes
@login_required(login_url='/')
def STAFF_NOTES(request):
    staff = Staff.objects.get(admin=request.user)
    notes = Note.objects.filter(user=staff.admin)
    context = {
        'notes': notes,
    }
    return render(request, 'Staff/notes.html', context)

@login_required(login_url='/')
def STAFF_CREATE_NOTE(request):
    if request.method == "POST":
        title = request.POST.get('title')
        content = request.POST.get('content')
        if title and content:
            note = Note(user=request.user, title=title, content=content)
            note.save()
            messages.success(request, 'Note created successfully!')
            return redirect('staff_notes')
        else:
            messages.error(request, 'Title and content are required.')
    return render(request, 'Staff/create_note.html')

@login_required(login_url='/')
def STAFF_EDIT_NOTE(request, note_id):
    try:
        note = Note.objects.get(id=note_id, user=request.user)
    except Note.DoesNotExist:
        messages.error(request, 'Note not found.')
        return redirect('staff_notes')
    if request.method == "POST":
        title = request.POST.get('title')
        content = request.POST.get('content')
        if title and content:
            note.title = title
            note.content = content
            note.save()
            messages.success(request, 'Note updated successfully!')
            return redirect('staff_notes')
        else:
            messages.error(request, 'Title and content are required.')
    context = {
        'note': note,
    }
    return render(request, 'Staff/edit_note.html', context)

@login_required(login_url='/')
def STAFF_DELETE_NOTE(request, note_id):
    try:
        note = Note.objects.get(id=note_id, user=request.user)
        note.delete()
        messages.success(request, 'Note deleted successfully!')
    except Note.DoesNotExist:
        messages.error(request, 'Note not found.')
    return redirect('staff_notes')